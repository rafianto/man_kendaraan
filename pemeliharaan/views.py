from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse  # <--- TAMBAHKAN BARIS INI
from django.db.models import Sum, Count, Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from xhtml2pdf import pisa
from django.template.loader import get_template

from .models import (
    Kendaraan, Bengkel, SparePart, JasaServis,
    Pemeliharaan, DetailSparePart, DetailJasaServis, KonsumsiBensin # TAMBAHKAN KonsumsiBensin
)
from .forms import (
    KendaraanForm, BengkelForm, SparePartForm, JasaServisForm,
    PemeliharaanForm, DetailSparePartForm, DetailJasaServisForm, KonsumsiBensinForm # TAMBAHKAN KonsumsiBensinForm
)


# ============================================================
# DASHBOARD
# ============================================================
@login_required
def dashboard(request):
    today = date.today()
    awal_minggu = today - timedelta(days=today.weekday())
    akhir_minggu = awal_minggu + timedelta(days=6)
    awal_bulan = today.replace(day=1)

    total_kendaraan = Kendaraan.objects.filter(aktif=True).count()
    total_kendaraan_mobil = Kendaraan.objects.filter(aktif=True, jenis='mobil').count()
    total_kendaraan_motor = Kendaraan.objects.filter(aktif=True, jenis='motor').count()
    total_kendaraan_sepeda = Kendaraan.objects.filter(aktif=True, jenis='sepeda_listrik').count()

    # === HARI INI ===
    servis_hari_ini = Pemeliharaan.objects.filter(tanggal_servis=today)
    total_hari_ini = servis_hari_ini.aggregate(t=Sum('total_biaya'))['t'] or 0
    sp_hari_ini = servis_hari_ini.aggregate(t=Sum('total_biaya_sparepart'))['t'] or 0
    js_hari_ini = servis_hari_ini.aggregate(t=Sum('total_biaya_jasa'))['t'] or 0
    
    # Bensin Hari Ini
    bensin_hari_ini = KonsumsiBensin.objects.filter(tanggal=today)
    total_bensin_hari = bensin_hari_ini.aggregate(t=Sum('total_biaya'))['t'] or 0

    # === MINGGU INI ===
    servis_minggu_ini = Pemeliharaan.objects.filter(tanggal_servis__range=[awal_minggu, akhir_minggu])
    total_minggu_ini = servis_minggu_ini.aggregate(t=Sum('total_biaya'))['t'] or 0
    sp_minggu_ini = servis_minggu_ini.aggregate(t=Sum('total_biaya_sparepart'))['t'] or 0
    js_minggu_ini = servis_minggu_ini.aggregate(t=Sum('total_biaya_jasa'))['t'] or 0
    
    # Bensin Minggu Ini
    bensin_minggu_ini = KonsumsiBensin.objects.filter(tanggal__range=[awal_minggu, akhir_minggu])
    total_bensin_minggu = bensin_minggu_ini.aggregate(t=Sum('total_biaya'))['t'] or 0

    # === BULAN INI ===
    servis_bulan_ini = Pemeliharaan.objects.filter(tanggal_servis__gte=awal_bulan)
    total_bulan_ini = servis_bulan_ini.aggregate(t=Sum('total_biaya'))['t'] or 0
    sp_bulan_ini = servis_bulan_ini.aggregate(t=Sum('total_biaya_sparepart'))['t'] or 0
    js_bulan_ini = servis_bulan_ini.aggregate(t=Sum('total_biaya_jasa'))['t'] or 0
    
    # Bensin Bulan Ini
    bensin_bulan_ini = KonsumsiBensin.objects.filter(tanggal__gte=awal_bulan)
    total_bensin_bulan = bensin_bulan_ini.aggregate(t=Sum('total_biaya'))['t'] or 0
    liter_bulan_ini = bensin_bulan_ini.aggregate(t=Sum('liter'))['t'] or 0

    # Grand Total Bulan Ini (Servis + Bensin)
    grand_total_bulan = total_bulan_ini + total_bensin_bulan

    # === PER KENDARAAN BULAN INI ===
    # === PER KENDARAAN BULAN INI ===
    kendaraan_biaya = []
    for k in Kendaraan.objects.filter(aktif=True):
        total_k = servis_bulan_ini.filter(kendaraan=k).aggregate(t=Sum('total_biaya'))['t'] or 0
        total_bensin_k = bensin_bulan_ini.filter(kendaraan=k).aggregate(t=Sum('total_biaya'))['t'] or 0
        liter_bensin_k = bensin_bulan_ini.filter(kendaraan=k).aggregate(t=Sum('liter'))['t'] or 0 # TAMBAHKAN INI
        combined_total = total_k + total_bensin_k
        if combined_total > 0:
            kendaraan_biaya.append({
                'kendaraan': k,
                'total_servis': total_k,
                'total_bensin': total_bensin_k,
                'liter_bensin': liter_bensin_k, # TAMBAHKAN INI
                'total_biaya': combined_total,
            })
    kendaraan_biaya.sort(key=lambda x: x['total_biaya'], reverse=True)

    # === SERVIS TERBARU ===
    servis_terbaru = Pemeliharaan.objects.all()[:5]
    bensin_terbaru = KonsumsiBensin.objects.all()[:5]

    context = {
        'total_kendaraan': total_kendaraan,
        'total_kendaraan_mobil': total_kendaraan_mobil,
        'total_kendaraan_motor': total_kendaraan_motor,
        'total_kendaraan_sepeda': total_kendaraan_sepeda,
        
        'total_hari_ini': total_hari_ini,
        'total_bensin_hari': total_bensin_hari,
        'sp_hari_ini': sp_hari_ini,
        'js_hari_ini': js_hari_ini,
        'jumlah_servis_hari': servis_hari_ini.count(),
        
        'total_minggu_ini': total_minggu_ini,
        'total_bensin_minggu': total_bensin_minggu,
        'sp_minggu_ini': sp_minggu_ini,
        'js_minggu_ini': js_minggu_ini,
        'jumlah_servis_minggu': servis_minggu_ini.count(),
        
        'total_bulan_ini': total_bulan_ini,
        'total_bensin_bulan': total_bensin_bulan,
        'liter_bulan_ini': liter_bulan_ini,
        'grand_total_bulan': grand_total_bulan,
        'sp_bulan_ini': sp_bulan_ini,
        'js_bulan_ini': js_bulan_ini,
        'jumlah_servis_bulan': servis_bulan_ini.count(),
        
        'kendaraan_biaya': kendaraan_biaya,
        'servis_terbaru': servis_terbaru,
        'bensin_terbaru': bensin_terbaru,
        'today': today,
        'awal_minggu': awal_minggu,
        'akhir_minggu': akhir_minggu,
        'awal_bulan': awal_bulan,
    }
    return render(request, 'pemeliharaan/dashboard.html', context)


# ============================================================
# KENDARAAN CRUD
# ============================================================
def kendaraan_list(request):
    queryset = Kendaraan.objects.all()
    jenis_filter = request.GET.get('jenis', '')
    search = request.GET.get('q', '')

    if jenis_filter:
        queryset = queryset.filter(jenis=jenis_filter)
    if search:
        queryset = queryset.filter(
            Q(nomor_polisi__icontains=search) |
            Q(merk__icontains=search) |
            Q(tipe__icontains=search) |
            Q(pemilik_pengguna__icontains=search)
        )

    context = {
        'object_list': queryset,
        'jenis_filter': jenis_filter,
        'search': search,
        'jenis_choices': JENIS_KENDARAAN if 'JENIS_KENDARAAN' in dir() else [
            ('mobil', 'Mobil'), ('motor', 'Motor'), ('sepeda_listrik', 'Sepeda Listrik')
        ],
    }
    return render(request, 'pemeliharaan/kendaraan_list.html', context)

@login_required
def kendaraan_create(request):
    if request.method == 'POST':
        form = KendaraanForm(request.POST, request.FILES) # TAMBAHKAN request.FILES
        if form.is_valid():
            form.save()
            messages.success(request, 'Data kendaraan berhasil ditambahkan.')
            return redirect('pemeliharaan:kendaraan_list')
    else:
        form = KendaraanForm()
    return render(request, 'pemeliharaan/kendaraan_form.html', {'form': form, 'title': 'Tambah Kendaraan'})


@login_required
def kendaraan_update(request, pk):
    obj = get_object_or_404(Kendaraan, pk=pk)
    if request.method == 'POST':
        form = KendaraanForm(request.POST, request.FILES, instance=obj) # TAMBAHKAN request.FILES
        if form.is_valid():
            form.save()
            messages.success(request, 'Data kendaraan berhasil diperbarui.')
            return redirect('pemeliharaan:kendaraan_list')
    else:
        form = KendaraanForm(instance=obj)
    return render(request, 'pemeliharaan/kendaraan_form.html', {'form': form, 'title': 'Edit Kendaraan'})

def kendaraan_delete(request, pk):
    obj = get_object_or_404(Kendaraan, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Data kendaraan berhasil dihapus.')
        return redirect('pemeliharaan:kendaraan_list')
    return render(request, 'pemeliharaan/kendaraan_form.html', {'object': obj})


@login_required
def kendaraan_detail(request, pk):
    obj = get_object_or_404(Kendaraan, pk=pk)
    servis_list = Pemeliharaan.objects.filter(kendaraan=obj)
    bensin_list = KonsumsiBensin.objects.filter(kendaraan=obj).order_by('-tanggal') # TAMBAHKAN INI
    
    today = date.today()
    awal_bulan = today.replace(day=1)

    total_semua = servis_list.aggregate(t=Sum('total_biaya'))['t'] or 0
    total_bulan = servis_list.filter(tanggal_servis__gte=awal_bulan).aggregate(
        t=Sum('total_biaya')
    )['t'] or 0
    total_sp = servis_list.aggregate(t=Sum('total_biaya_sparepart'))['t'] or 0
    total_js = servis_list.aggregate(t=Sum('total_biaya_jasa'))['t'] or 0

    # PERHITUNGAN BENSIN (BARU)
    total_bensin = bensin_list.aggregate(t=Sum('total_biaya'))['t'] or 0
    total_liter = bensin_list.aggregate(t=Sum('liter'))['t'] or 0
    jumlah_isi = bensin_list.count()
    grand_total = total_semua + total_bensin

    context = {
        'kendaraan': obj,
        'servis_list': servis_list,
        'bensin_list': bensin_list, # TAMBAHKAN INI
        'total_semua': total_semua,
        'total_bulan': total_bulan,
        'total_sp': total_sp,
        'total_js': total_js,
        'jumlah_servis': servis_list.count(),
        'total_bensin': total_bensin, # TAMBAHKAN INI
        'total_liter': total_liter,   # TAMBAHKAN INI
        'jumlah_isi': jumlah_isi,     # TAMBAHKAN INI
        'grand_total': grand_total,   # TAMBAHKAN INI
    }
    return render(request, 'pemeliharaan/kendaraan_detail.html', context)

# ============================================================
# BENGKEL CRUD
# ============================================================
def bengkel_list(request):
    search = request.GET.get('q', '')
    queryset = Bengkel.objects.all()
    if search:
        queryset = queryset.filter(
            Q(nama_bengkel__icontains=search) |
            Q(pemilik__icontains=search) |
            Q(kota__icontains=search)
        )
    return render(request, 'pemeliharaan/bengkel_list.html', {
        'object_list': queryset, 'search': search
    })


def bengkel_create(request):
    if request.method == 'POST':
        form = BengkelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data bengkel berhasil ditambahkan.')
            return redirect('pemeliharaan:bengkel_list')
    else:
        form = BengkelForm()
    return render(request, 'pemeliharaan/bengkel_form.html', {'form': form, 'title': 'Tambah Bengkel'})


def bengkel_update(request, pk):
    obj = get_object_or_404(Bengkel, pk=pk)
    if request.method == 'POST':
        form = BengkelForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data bengkel berhasil diperbarui.')
            return redirect('pemeliharaan:bengkel_list')
    else:
        form = BengkelForm(instance=obj)
    return render(request, 'pemeliharaan/bengkel_form.html', {'form': form, 'title': 'Edit Bengkel'})


def bengkel_delete(request, pk):
    obj = get_object_or_404(Bengkel, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Data bengkel berhasil dihapus.')
        return redirect('pemeliharaan:bengkel_list')
    return redirect('pemeliharaan:bengkel_list')


# ============================================================
# SPAREPART CRUD
# ============================================================
def sparepart_list(request):
    search = request.GET.get('q', '')
    queryset = SparePart.objects.all()
    if search:
        queryset = queryset.filter(
            Q(kode__icontains=search) | Q(nama_sparepart__icontains=search)
        )
    return render(request, 'pemeliharaan/sparepart_list.html', {
        'object_list': queryset, 'search': search
    })


def sparepart_create(request):
    if request.method == 'POST':
        form = SparePartForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data spare part berhasil ditambahkan.')
            return redirect('pemeliharaan:sparepart_list')
    else:
        form = SparePartForm()
    return render(request, 'pemeliharaan/sparepart_form.html', {'form': form, 'title': 'Tambah Spare Part'})


def sparepart_update(request, pk):
    obj = get_object_or_404(SparePart, pk=pk)
    if request.method == 'POST':
        form = SparePartForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data spare part berhasil diperbarui.')
            return redirect('pemeliharaan:sparepart_list')
    else:
        form = SparePartForm(instance=obj)
    return render(request, 'pemeliharaan/sparepart_form.html', {'form': form, 'title': 'Edit Spare Part'})


def sparepart_delete(request, pk):
    obj = get_object_or_404(SparePart, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Data spare part berhasil dihapus.')
    return redirect('pemeliharaan:sparepart_list')


# ============================================================
# JASA SERVIS CRUD
# ============================================================
def jasaservis_list(request):
    search = request.GET.get('q', '')
    queryset = JasaServis.objects.all()
    if search:
        queryset = queryset.filter(
            Q(kode__icontains=search) | Q(nama_jasa__icontains=search)
        )
    return render(request, 'pemeliharaan/jasaservis_list.html', {
        'object_list': queryset, 'search': search
    })


def jasaservis_create(request):
    if request.method == 'POST':
        form = JasaServisForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data jasa servis berhasil ditambahkan.')
            return redirect('pemeliharaan:jasaservis_list')
    else:
        form = JasaServisForm()
    return render(request, 'pemeliharaan/jasaservis_form.html', {'form': form, 'title': 'Tambah Jasa Servis'})


def jasaservis_update(request, pk):
    obj = get_object_or_404(JasaServis, pk=pk)
    if request.method == 'POST':
        form = JasaServisForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data jasa servis berhasil diperbarui.')
            return redirect('pemeliharaan:jasaservis_list')
    else:
        form = JasaServisForm(instance=obj)
    return render(request, 'pemeliharaan/jasaservis_form.html', {'form': form, 'title': 'Edit Jasa Servis'})


def jasaservis_delete(request, pk):
    obj = get_object_or_404(JasaServis, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Data jasa servis berhasil dihapus.')
    return redirect('pemeliharaan:jasaservis_list')


# ============================================================
# PEMELIHARAAN CRUD
# ============================================================
def pemeliharaan_list(request):
    queryset = Pemeliharaan.objects.select_related('kendaraan', 'bengkel').all()
    search = request.GET.get('q', '')
    start = request.GET.get('start', '')
    end = request.GET.get('end', '')
    jenis = request.GET.get('jenis', '')

    if search:
        queryset = queryset.filter(
            Q(no_servis__icontains=search) |
            Q(kendaraan__nomor_polisi__icontains=search) |
            Q(kendaraan__merk__icontains=search)
        )
    if start:
        queryset = queryset.filter(tanggal_servis__gte=start)
    if end:
        queryset = queryset.filter(tanggal_servis__lte=end)
    if jenis:
        queryset = queryset.filter(kendaraan__jenis=jenis)

    context = {
        'object_list': queryset,
        'search': search,
        'start': start,
        'end': end,
        'jenis': jenis,
    }
    return render(request, 'pemeliharaan/pemeliharaan_list.html', context)


def pemeliharaan_create(request):
    if request.method == 'POST':
        form = PemeliharaanForm(request.POST)
        if form.is_valid():
            pem = form.save()
            messages.success(request, f'Servis {pem.no_servis} berhasil dibuat. Tambahkan detail spare part dan jasa.')
            return redirect('pemeliharaan:pemeliharaan_detail', pk=pem.pk)
    else:
        form = PemeliharaanForm()
    return render(request, 'pemeliharaan/pemeliharaan_form.html', {'form': form, 'title': 'Tambah Pemeliharaan'})


def pemeliharaan_update(request, pk):
    obj = get_object_or_404(Pemeliharaan, pk=pk)
    if request.method == 'POST':
        form = PemeliharaanForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data pemeliharaan berhasil diperbarui.')
            return redirect('pemeliharaan:pemeliharaan_detail', pk=obj.pk)
    else:
        form = PemeliharaanForm(instance=obj)
    return render(request, 'pemeliharaan/pemeliharaan_form.html', {'form': form, 'title': 'Edit Pemeliharaan'})


def pemeliharaan_delete(request, pk):
    obj = get_object_or_404(Pemeliharaan, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Data pemeliharaan berhasil dihapus.')
        return redirect('pemeliharaan:pemeliharaan_list')
    return redirect('pemeliharaan:pemeliharaan_list')


def pemeliharaan_detail(request, pk):
    obj = get_object_or_404(Pemeliharaan, pk=pk)
    detail_sp = DetailSparePart.objects.filter(pemeliharaan=obj).select_related('sparepart')
    detail_js = DetailJasaServis.objects.filter(pemeliharaan=obj).select_related('jasa_servis')

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'sparepart':
            sp_form = DetailSparePartForm(request.POST)
            js_form = DetailJasaServisForm()
            if sp_form.is_valid():
                detail = sp_form.save(commit=False)
                detail.pemeliharaan = obj
                detail.save()
                messages.success(request, 'Spare part berhasil ditambahkan.')
                return redirect('pemeliharaan:pemeliharaan_detail', pk=obj.pk)

        elif form_type == 'jasaservis':
            sp_form = DetailSparePartForm()
            js_form = DetailJasaServisForm(request.POST)
            if js_form.is_valid():
                detail = js_form.save(commit=False)
                detail.pemeliharaan = obj
                detail.save()
                messages.success(request, 'Jasa servis berhasil ditambahkan.')
                return redirect('pemeliharaan:pemeliharaan_detail', pk=obj.pk)
    else:
        sp_form = DetailSparePartForm()
        js_form = DetailJasaServisForm()

    # Refresh object untuk mendapat total terbaru
    obj.refresh_from_db()

    # Data harga untuk auto-fill JavaScript
    sparepart_data = {str(sp.pk): float(sp.harga) for sp in SparePart.objects.all()}
    jasaservis_data = {str(js.pk): float(js.biaya) for js in JasaServis.objects.all()}

    context = {
        'pemeliharaan': obj,
        'detail_sp': detail_sp,
        'detail_js': detail_js,
        'sp_form': sp_form,
        'js_form': js_form,
        'sparepart_data': sparepart_data,
        'jasaservis_data': jasaservis_data,
    }
    return render(request, 'pemeliharaan/pemeliharaan_detail.html', context)


def hapus_detail_sparepart(request, pk, detail_pk):
    detail = get_object_or_404(DetailSparePart, pk=detail_pk, pemeliharaan__pk=pk)
    if request.method == 'POST':
        detail.delete()
        messages.success(request, 'Detail spare part berhasil dihapus.')
    return redirect('pemeliharaan:pemeliharaan_detail', pk=pk)


def hapus_detail_jasaservis(request, pk, detail_pk):
    detail = get_object_or_404(DetailJasaServis, pk=detail_pk, pemeliharaan__pk=pk)
    if request.method == 'POST':
        detail.delete()
        messages.success(request, 'Detail jasa servis berhasil dihapus.')
    return redirect('pemeliharaan:pemeliharaan_detail', pk=pk)


# ============================================================
# LAPORAN
@login_required
@login_required
def laporan(request):
    today = date.today()
    periode = request.GET.get('periode', 'bulanan')
    kendaraan_id = request.GET.get('kendaraan', '')
    custom_start = request.GET.get('start_date', '')
    custom_end = request.GET.get('end_date', '')

    # Tentukan range tanggal (KODE INI TETAP SAMA SEPERTI SEBELUMNYA)
    if periode == 'harian':
        start_date = today
        end_date = today
    elif periode == 'mingguan':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif periode == 'bulanan':
        start_date = today.replace(day=1)
        end_date = today
    elif periode == 'custom':
        start_date = today.replace(day=1)
        end_date = today
        if custom_start:
            try: start_date = date.fromisoformat(custom_start)
            except ValueError: pass
        if custom_end:
            try: end_date = date.fromisoformat(custom_end)
            except ValueError: pass
    else:
        start_date = today.replace(day=1)
        end_date = today

    # Filter Pemeliharaan (SAMPAI SINI TETAP SAMA)

    queryset = Pemeliharaan.objects.filter(tanggal_servis__range=[start_date, end_date]).select_related('kendaraan', 'bengkel')
    if kendaraan_id:
        queryset = queryset.filter(kendaraan_id=kendaraan_id)

    total_sparepart = queryset.aggregate(t=Sum('total_biaya_sparepart'))['t'] or 0
    total_jasa = queryset.aggregate(t=Sum('total_biaya_jasa'))['t'] or 0
    total_biaya = queryset.aggregate(t=Sum('total_biaya'))['t'] or 0
    jumlah_servis = queryset.count()

    kendaraan_breakdown = queryset.values('kendaraan__pk', 'kendaraan__nomor_polisi', 'kendaraan__merk', 'kendaraan__tipe', 'kendaraan__jenis').annotate(
        total_sparepart=Sum('total_biaya_sparepart'), total_jasa=Sum('total_biaya_jasa'), total_biaya=Sum('total_biaya'), jumlah_servis=Count('id')
    ).order_by('-total_biaya')

    bengkel_breakdown = queryset.values('bengkel__pk', 'bengkel__nama_bengkel', 'bengkel__kota').annotate(
        total_sparepart=Sum('total_biaya_sparepart'), total_jasa=Sum('total_biaya_jasa'), total_biaya=Sum('total_biaya'), jumlah_servis=Count('id')
    ).order_by('-total_biaya')

    # === TAMBAHAN PERHITUNGAN BENSIN DI LAPORAN ===
    queryset_bensin = KonsumsiBensin.objects.filter(tanggal__range=[start_date, end_date]).select_related('kendaraan')
    if kendaraan_id:
        queryset_bensin = queryset_bensin.filter(kendaraan_id=kendaraan_id)

    total_bensin = queryset_bensin.aggregate(t=Sum('total_biaya'))['t'] or 0
    total_liter = queryset_bensin.aggregate(t=Sum('liter'))['t'] or 0
    jumlah_pengisian = queryset_bensin.count()

    bensin_breakdown = queryset_bensin.values('kendaraan__pk', 'kendaraan__nomor_polisi', 'kendaraan__merk', 'kendaraan__jenis').annotate(
        total_bensin=Sum('total_biaya'), total_liter=Sum('liter'), jumlah_isi=Count('id')
    ).order_by('-total_bensin')

    # Grand Total (Pemeliharaan + Bensin)
    grand_total_biaya = total_biaya + total_bensin

    context = {
        'periode': periode,
        'start_date': start_date,
        'end_date': end_date,
        'total_sparepart': total_sparepart,
        'total_jasa': total_jasa,
        'total_biaya': total_biaya,
        'jumlah_servis': jumlah_servis,
        'kendaraan_breakdown': kendaraan_breakdown,
        'bengkel_breakdown': bengkel_breakdown,
        'pemeliharaan_list': queryset.order_by('-tanggal_servis'),
        
        # Data Bensin
        'total_bensin': total_bensin,
        'total_liter': total_liter,
        'jumlah_pengisian': jumlah_pengisian,
        'bensin_breakdown': bensin_breakdown,
        'bensin_list': queryset_bensin.order_by('-tanggal'),
        'grand_total_biaya': grand_total_biaya,
        
        'kendaraan_list': Kendaraan.objects.filter(aktif=True),
        'selected_kendaraan': kendaraan_id,
        'custom_start': custom_start or start_date.isoformat(),
        'custom_end': custom_end or end_date.isoformat(),
    }
    return render(request, 'pemeliharaan/laporan.html', context)

@login_required
def export_laporan_excel(request):
    # Ambil parameter filter yang sama persis dengan halaman laporan
    today = date.today()
    periode = request.GET.get('periode', 'bulanan')
    kendaraan_id = request.GET.get('kendaraan', '')
    custom_start = request.GET.get('start_date', '')
    custom_end = request.GET.get('end_date', '')

    if periode == 'harian':
        start_date = today
        end_date = today
    elif periode == 'mingguan':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif periode == 'bulanan':
        start_date = today.replace(day=1)
        end_date = today
    elif periode == 'custom':
        start_date = today.replace(day=1)
        end_date = today
        if custom_start:
            try: start_date = date.fromisoformat(custom_start)
            except ValueError: pass
        if custom_end:
            try: end_date = date.fromisoformat(custom_end)
            except ValueError: pass
    else:
        start_date = today.replace(day=1)
        end_date = today

    # === FILTER PEMELIHARAAN ===
    queryset = Pemeliharaan.objects.filter(tanggal_servis__range=[start_date, end_date]).select_related('kendaraan', 'bengkel')
    if kendaraan_id:
        queryset = queryset.filter(kendaraan_id=kendaraan_id)

    total_sparepart = queryset.aggregate(t=Sum('total_biaya_sparepart'))['t'] or 0
    total_jasa = queryset.aggregate(t=Sum('total_biaya_jasa'))['t'] or 0
    total_biaya_servis = queryset.aggregate(t=Sum('total_biaya'))['t'] or 0
    jumlah_servis = queryset.count()

    kendaraan_breakdown = queryset.values(
        'kendaraan__nomor_polisi', 'kendaraan__merk', 'kendaraan__jenis'
    ).annotate(
        total_sparepart=Sum('total_biaya_sparepart'),
        total_jasa=Sum('total_biaya_jasa'),
        total_biaya=Sum('total_biaya'),
        jumlah_servis=Count('id')
    ).order_by('-total_biaya')

    bengkel_breakdown = queryset.values(
        'bengkel__nama_bengkel', 'bengkel__kota'
    ).annotate(
        total_sparepart=Sum('total_biaya_sparepart'),
        total_jasa=Sum('total_biaya_jasa'),
        total_biaya=Sum('total_biaya'),
        jumlah_servis=Count('id')
    ).order_by('-total_biaya')

    # === FILTER BENSIN (BARU) ===
    queryset_bensin = KonsumsiBensin.objects.filter(tanggal__range=[start_date, end_date]).select_related('kendaraan')
    if kendaraan_id:
        queryset_bensin = queryset_bensin.filter(kendaraan_id=kendaraan_id)

    total_bensin = queryset_bensin.aggregate(t=Sum('total_biaya'))['t'] or 0
    total_liter = queryset_bensin.aggregate(t=Sum('liter'))['t'] or 0
    jumlah_pengisian = queryset_bensin.count()

    bensin_breakdown = queryset_bensin.values(
        'kendaraan__nomor_polisi', 'kendaraan__merk', 'kendaraan__jenis'
    ).annotate(
        total_bensin=Sum('total_biaya'),
        total_liter=Sum('liter'),
        jumlah_isi=Count('id')
    ).order_by('-total_bensin')

    # Grand Total
    grand_total_biaya = total_biaya_servis + total_bensin

    # === MEMBUAT FILE EXCEL ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Pemeliharaan"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    bensin_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    grand_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="1E293B")
    bold_font = Font(bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    idr_format = '#,##0'
    liter_format = '#,##0.00'

    # Set lebar kolom
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20

    # 1. Header Laporan
    ws.merge_cells('A1:H1')
    ws['A1'] = "LAPORAN BIAYA PEMELIHARAAN & BENSIN KENDARAAN"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Periode: {start_date.strftime('%d %B %Y')} s/d {end_date.strftime('%d %B %Y')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    # 2. Ringkasan Biaya
    row = 4
    ws.cell(row=row, column=1, value="RINGKASAN BIAYA").font = subtitle_font
    row += 1
    summary_headers = ["Keterangan", "Total"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border

    row += 1
    summary_data = [
        ("Total Spare Part", total_sparepart),
        ("Total Jasa Servis", total_jasa),
        ("Total Pemeliharaan", total_biaya_servis),
        ("", ""),
        ("Total Bensin", total_bensin),
        ("Total Liter Bensin", total_liter),
        ("Jumlah Pengisian Bensin", jumlah_pengisian),
        ("", ""),
        ("GRAND TOTAL (Servis + Bensin)", grand_total_biaya),
    ]
    for label, val in summary_data:
        if not label: row += 1; continue
        ws.cell(row=row, column=1, value=label).border = thin_border
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.border = thin_border
        
        if label in ["Total Liter Bensin"]:
            cell_val.number_format = liter_format
        elif label != "Jumlah Pengisian Bensin":
            cell_val.number_format = idr_format
            
        if "GRAND TOTAL" in label:
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            cell_val.font = Font(bold=True, size=12)
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = grand_fill
                ws.cell(row=row, column=col).font = Font(bold=True, size=12, color="FFFFFF")
        row += 1

    # 3. Biaya Pemeliharaan per Kendaraan
    row += 1
    ws.cell(row=row, column=1, value="BIAYA PEMELIHARAAN PER KENDARAAN").font = subtitle_font
    row += 1
    kend_headers = ["No", "Nopol", "Merk", "Jenis", "Spare Part", "Jasa", "Total Biaya"]
    for col_num, header in enumerate(kend_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

    row += 1
    for idx, item in enumerate(kendaraan_breakdown, 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=item['kendaraan__nomor_polisi']).border = thin_border
        ws.cell(row=row, column=3, value=item['kendaraan__merk']).border = thin_border
        ws.cell(row=row, column=4, value=item['kendaraan__jenis']).border = thin_border
        c1 = ws.cell(row=row, column=5, value=item['total_sparepart']); c1.number_format = idr_format; c1.border = thin_border
        c2 = ws.cell(row=row, column=6, value=item['total_jasa']); c2.number_format = idr_format; c2.border = thin_border
        c3 = ws.cell(row=row, column=7, value=item['total_biaya']); c3.number_format = idr_format; c3.border = thin_border; c3.font = bold_font
        row += 1

    # 4. Biaya Bensin per Kendaraan (BARU)
    row += 1
    ws.cell(row=row, column=1, value="BIAYA BENSIN PER KENDARAAN").font = subtitle_font
    row += 1
    bensin_kend_headers = ["No", "Nopol", "Merk", "Jenis", "Jumlah Isi", "Total Liter", "Total Biaya"]
    for col_num, header in enumerate(bensin_kend_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font; cell.fill = bensin_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

    row += 1
    for idx, item in enumerate(bensin_breakdown, 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=item['kendaraan__nomor_polisi']).border = thin_border
        ws.cell(row=row, column=3, value=item['kendaraan__merk']).border = thin_border
        ws.cell(row=row, column=4, value=item['kendaraan__jenis']).border = thin_border
        ws.cell(row=row, column=5, value=item['jumlah_isi']).border = thin_border
        c1 = ws.cell(row=row, column=6, value=item['total_liter']); c1.number_format = liter_format; c1.border = thin_border
        c2 = ws.cell(row=row, column=7, value=item['total_bensin']); c2.number_format = idr_format; c2.border = thin_border; c2.font = bold_font
        row += 1

    # 5. Detail Transaksi Pemeliharaan
    row += 1
    ws.cell(row=row, column=1, value="DETAIL TRANSAKSI PEMELIHARAAN").font = subtitle_font
    row += 1
    trans_headers = ["No", "Tanggal", "No Servis", "Kendaraan", "Bengkel", "Spare Part", "Jasa", "Total"]
    for col_num, header in enumerate(trans_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

    row += 1
    for idx, s in enumerate(queryset.order_by('-tanggal_servis'), 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=s.tanggal_servis.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(row=row, column=3, value=s.no_servis).border = thin_border
        ws.cell(row=row, column=4, value=s.kendaraan.nomor_polisi).border = thin_border
        ws.cell(row=row, column=5, value=s.bengkel.nama_bengkel).border = thin_border
        c1 = ws.cell(row=row, column=6, value=s.total_biaya_sparepart); c1.number_format = idr_format; c1.border = thin_border
        c2 = ws.cell(row=row, column=7, value=s.total_biaya_jasa); c2.number_format = idr_format; c2.border = thin_border
        c3 = ws.cell(row=row, column=8, value=s.total_biaya); c3.number_format = idr_format; c3.border = thin_border; c3.font = Font(bold=True, color="3B82F6")
        row += 1

    # 6. Detail Transaksi Bensin (BARU)
    row += 1
    ws.cell(row=row, column=1, value="DETAIL PENGISIAN BENSIN").font = subtitle_font
    row += 1
    bensin_headers = ["No", "Tanggal", "Kendaraan", "KM", "SPBU", "Liter", "Harga/Liter", "Total Biaya"]
    for col_num, header in enumerate(bensin_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font; cell.fill = bensin_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

    row += 1
    for idx, b in enumerate(queryset_bensin.order_by('-tanggal'), 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=b.tanggal.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(row=row, column=3, value=b.kendaraan.nomor_polisi).border = thin_border
        ws.cell(row=row, column=4, value=b.km_pengisian).border = thin_border
        ws.cell(row=row, column=5, value=b.spbu).border = thin_border
        c1 = ws.cell(row=row, column=6, value=b.liter); c1.number_format = liter_format; c1.border = thin_border
        c2 = ws.cell(row=row, column=7, value=b.harga_per_liter); c2.number_format = idr_format; c2.border = thin_border
        c3 = ws.cell(row=row, column=8, value=b.total_biaya); c3.number_format = idr_format; c3.border = thin_border; c3.font = Font(bold=True, color="F59E0B")
        row += 1

    # Return Excel File
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Laporan_Pemeliharaan_Bensin.xlsx"'
    wb.save(response)
    return response

@login_required
def export_kuitansi_excel(request, pk):
    # Ambil data pemeliharaan berdasarkan pk
    pemeliharaan = get_object_or_404(Pemeliharaan, pk=pk)
    detail_sp = DetailSparePart.objects.filter(pemeliharaan=pemeliharaan).select_related('sparepart')
    detail_js = DetailJasaServis.objects.filter(pemeliharaan=pemeliharaan).select_related('jasa_servis')
    
    # Refresh untuk memastikan total terbaru
    pemeliharaan.refresh_from_db()

    # === MEMBUAT FILE EXCEL KUITANSI ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Kuitansi {pemeliharaan.no_servis}"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_font = Font(bold=True, size=16, color="1E293B")
    subtitle_font = Font(bold=True, size=11, color="3B82F6")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='medium', color='1E293B'))
    idr_format = '#,##0'

    # Set lebar kolom
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20

    # ==========================================
    # HEADER KUITANSI
    # ==========================================
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "👁 OPTIC GADING FORCE"
    cell.font = Font(bold=True, size=20, color="3B82F6")
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = "Divisi Kendaraan"
    cell.font = Font(bold=True, size=11, color="1E293B")

    ws.merge_cells('A3:E3')
    cell = ws['A3']
    cell.value = "Ds. Guyangan RT 01 RW 03 No 6 Madureso, Kuwarasan | Telp: (021) 555-8899"
    cell.font = Font(size=9, color="64748B")

    # Garis pemisah
    for col in range(1, 6):
        ws.cell(row=4, column=col).border = bottom_border
    ws.row_dimensions[4].height = 8

    # Judul Kuitansi
    ws.merge_cells('A5:E5')
    cell = ws['A5']
    cell.value = "KUITANSI / INVOICE"
    cell.font = Font(bold=True, size=14, color="1E293B")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 30

    ws.merge_cells('A6:E6')
    cell = ws['A6']
    cell.value = "BUKTI PEMBAYARAN JASA & SPARE PART KENDARAAN"
    cell.font = Font(size=9, color="64748B")
    cell.alignment = Alignment(horizontal='center')

    # ==========================================
    # INFORMASI KUITANSI
    # ==========================================
    row = 8
    
    # Info Servis (Kolom A-C)
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="No. Servis:").font = bold_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=pemeliharaan.no_servis).font = Font(bold=True, size=11, color="3B82F6")
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Tanggal Servis:").font = normal_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=pemeliharaan.tanggal_servis.strftime('%d %B %Y')).font = normal_font
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="KM Saat Servis:").font = normal_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=f"{pemeliharaan.km_saat_servis:,} KM" if pemeliharaan.km_saat_servis else "-").font = normal_font

    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Kendaraan:").font = bold_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=f"{pemeliharaan.kendaraan.nomor_polisi} - {pemeliharaan.kendaraan.merk} {pemeliharaan.kendaraan.tipe}").font = normal_font
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Jenis:").font = normal_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=pemeliharaan.kendaraan.get_jenis_display()).font = normal_font
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Pemilik/Pengguna:").font = normal_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=pemeliharaan.kendaraan.pemilik_pengguna).font = normal_font
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Bengkel:").font = bold_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=pemeliharaan.bengkel.nama_bengkel).font = normal_font
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Alamat Bengkel:").font = normal_font
    ws.merge_cells(f'C{row}:E{row}')
    ws.cell(row=row, column=3, value=f"{pemeliharaan.bengkel.alamat}, {pemeliharaan.bengkel.kota}").font = normal_font

    # Garis pemisah
    row += 2
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = bottom_border
    ws.row_dimensions[row].height = 5

    # ==========================================
    # TABEL SPARE PART
    # ==========================================
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="A. SPARE PART YANG DIGANTI").font = subtitle_font
    row += 1

    sp_headers = ["No", "Nama Spare Part", "Jumlah", "Harga Satuan", "Subtotal"]
    for col_num, header in enumerate(sp_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row += 1
    if detail_sp.exists():
        for idx, d in enumerate(detail_sp, 1):
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=d.sparepart.nama_sparepart).border = thin_border
            ws.cell(row=row, column=3, value=d.jumlah).border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            c1 = ws.cell(row=row, column=4, value=d.harga_satuan)
            c1.number_format = idr_format; c1.border = thin_border; c1.alignment = Alignment(horizontal='right')
            
            c2 = ws.cell(row=row, column=5, value=d.subtotal)
            c2.number_format = idr_format; c2.border = thin_border; c2.alignment = Alignment(horizontal='right')
            c2.font = bold_font
            row += 1
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1, value="Tidak ada spare part yang diganti").font = Font(italic=True, color="94A3B8")
        row += 1

    # Subtotal Spare Part
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="Subtotal Spare Part").font = bold_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
    for col_num in range(1, 5):
        ws.cell(row=row, column=col_num).border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    
    cell_total_sp = ws.cell(row=row, column=5, value=pemeliharaan.total_biaya_sparepart)
    cell_total_sp.number_format = idr_format
    cell_total_sp.font = Font(bold=True, size=11)
    cell_total_sp.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    cell_total_sp.alignment = Alignment(horizontal='right')

    # ==========================================
    # TABEL JASA SERVIS
    # ==========================================
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="B. JASA SERVIS").font = subtitle_font
    row += 1

    # Header Jasa Servis (Perbaikan merge cell)
    ws.cell(row=row, column=1, value="No").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value="Nama Jasa Servis").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    
    # Beri border untuk area merge
    for c in range(3, 5):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = header_fill

    ws.cell(row=row, column=5, value="Biaya").font = header_font
    ws.cell(row=row, column=5).fill = header_fill
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')

    row += 1
    if detail_js.exists():
        for idx, d in enumerate(detail_js, 1):
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            ws.merge_cells(f'B{row}:D{row}')
            ws.cell(row=row, column=2, value=d.jasa_servis.nama_jasa).border = thin_border
            for c in range(3, 5):
                ws.cell(row=row, column=c).border = thin_border
            
            c1 = ws.cell(row=row, column=5, value=d.biaya)
            c1.number_format = idr_format; c1.border = thin_border; c1.alignment = Alignment(horizontal='right')
            c1.font = bold_font
            row += 1
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1, value="Tidak ada jasa servis").font = Font(italic=True, color="94A3B8")
        row += 1

    # Subtotal Jasa
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="Subtotal Jasa Servis").font = bold_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
    for col_num in range(1, 5):
        ws.cell(row=row, column=col_num).border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    
    cell_total_js = ws.cell(row=row, column=5, value=pemeliharaan.total_biaya_jasa)
    cell_total_js.number_format = idr_format
    cell_total_js.font = Font(bold=True, size=11)
    cell_total_js.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    cell_total_js.alignment = Alignment(horizontal='right')

    # ==========================================
    # GRAND TOTAL
    # ==========================================
    row += 2
    grand_total_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    ws.merge_cells(f'A{row}:D{row}')
    cell_label = ws.cell(row=row, column=1, value="GRAND TOTAL (Rp)")
    cell_label.font = Font(bold=True, size=13, color="FFFFFF")
    cell_label.fill = grand_total_fill
    cell_label.alignment = Alignment(horizontal='right', vertical='center')
    for col_num in range(1, 5):
        ws.cell(row=row, column=col_num).fill = grand_total_fill
        ws.cell(row=row, column=col_num).border = thin_border
    
    cell_grand = ws.cell(row=row, column=5, value=pemeliharaan.total_biaya)
    cell_grand.number_format = idr_format
    cell_grand.font = Font(bold=True, size=13, color="FFFFFF")
    cell_grand.fill = grand_total_fill
    cell_grand.alignment = Alignment(horizontal='right', vertical='center')
    cell_grand.border = thin_border
    ws.row_dimensions[row].height = 30

    # ==========================================
    # TERBILANG
    # ==========================================
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    
    def terbilang(n):
        if n == 0: return "Nol"
        angka = ["", "Satu", "Dua", "Tiga", "Empat", "Lima", "Enam", "Tujuh", "Delapan", "Sembilan", "Sepuluh", "Sebelas"]
        if n < 12:
            return angka[n]
        elif n < 20:
            return terbilang(n - 10) + " Belas"
        elif n < 100:
            return terbilang(n // 10) + " Puluh" + (" " + terbilang(n % 10) if n % 10 else "")
        elif n < 200:
            return "Seratus" + (" " + terbilang(n - 100) if n - 100 else "")
        elif n < 1000:
            return terbilang(n // 100) + " Ratus" + (" " + terbilang(n % 100) if n % 100 else "")
        elif n < 2000:
            return "Seribu" + (" " + terbilang(n - 1000) if n - 1000 else "")
        elif n < 1000000:
            return terbilang(n // 1000) + " Ribu" + (" " + terbilang(n % 1000) if n % 1000 else "")
        elif n < 1000000000:
            return terbilang(n // 1000000) + " Juta" + (" " + terbilang(n % 1000000) if n % 1000000 else "")
        else:
            return str(n)
    
    try:
        total_int = int(pemeliharaan.total_biaya)
        bilang = terbilang(total_int) + " Rupiah"
    except:
        bilang = "-"
    
    ws.cell(row=row, column=1, value=f"Terbilang: {bilang}").font = Font(italic=True, bold=True, size=10, color="1E293B")

    # ==========================================
    # KETERANGAN & TANDA TANGAN
    # ==========================================
    row += 2
    if pemeliharaan.keterangan:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1, value=f"Keterangan: {pemeliharaan.keterangan}").font = Font(italic=True, size=10, color="64748B")
        row += 1

    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Penerima,").font = normal_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'D{row}:E{row}')
    ws.cell(row=row, column=4, value="Hormat Kami,").font = normal_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

    row += 5  # Ruang untuk tanda tangan
    
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="____________________").font = normal_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'D{row}:E{row}')
    ws.cell(row=row, column=4, value="____________________").font = normal_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Pemilik / Pengguna").font = Font(size=9, color="64748B")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'D{row}:E{row}')
    ws.cell(row=row, column=4, value="Admin Kendaraan").font = Font(size=9, color="64748B")
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

    # ==========================================
    # FOOTER
    # ==========================================
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="Kuitansi ini sah dan dibuat dengan sebenarnya.").font = Font(size=8, italic=True, color="94A3B8")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=f"Dicetak pada: {date.today().strftime('%d %B %Y')} | OptiCare Fleet Maintenance System").font = Font(size=8, italic=True, color="94A3B8")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # Print setup
    ws.print_area = f'A1:E{row}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Return Excel File
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Kuitansi_{pemeliharaan.no_servis}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# KONSUMSI BENSIN CRUD
# ============================================================
@login_required
def bensin_list(request):
    queryset = KonsumsiBensin.objects.select_related('kendaraan').all()
    search = request.GET.get('q', '')
    start = request.GET.get('start', '')
    end = request.GET.get('end', '')

    if search:
        queryset = queryset.filter(
            Q(kendaraan__nomor_polisi__icontains=search) |
            Q(spbu__icontains=search)
        )
    if start:
        queryset = queryset.filter(tanggal__gte=start)
    if end:
        queryset = queryset.filter(tanggal__lte=end)

    # Hitung total keseluruhan untuk data yang difilter
    total_keseluruhan = queryset.aggregate(t=Sum('total_biaya'))['t'] or 0
    total_liter = queryset.aggregate(t=Sum('liter'))['t'] or 0

    context = {
        'object_list': queryset,
        'search': search,
        'start': start,
        'end': end,
        'total_keseluruhan': total_keseluruhan,
        'total_liter': total_liter,
    }
    return render(request, 'pemeliharaan/bensin_list.html', context)


@login_required
def bensin_create(request):
    if request.method == 'POST':
        form = KonsumsiBensinForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data pengisian bensin berhasil ditambahkan.')
            return redirect('pemeliharaan:bensin_list')
    else:
        form = KonsumsiBensinForm()
    return render(request, 'pemeliharaan/bensin_form.html', {'form': form, 'title': 'Tambah Pengisian Bensin'})


@login_required
def bensin_update(request, pk):
    obj = get_object_or_404(KonsumsiBensin, pk=pk)
    if request.method == 'POST':
        form = KonsumsiBensinForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data pengisian bensin berhasil diperbarui.')
            return redirect('pemeliharaan:bensin_list')
    else:
        form = KonsumsiBensinForm(instance=obj)
    return render(request, 'pemeliharaan/bensin_form.html', {'form': form, 'title': 'Edit Pengisian Bensin'})


@login_required
def bensin_delete(request, pk):
    obj = get_object_or_404(KonsumsiBensin, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Data pengisian bensin berhasil dihapus.')
    return redirect('pemeliharaan:bensin_list')


@login_required
def export_struk_bensin_pdf(request, pk):
    # Ambil data bensin berdasarkan pk
    bensin = get_object_or_404(KonsumsiBensin, pk=pk)
    
    # Siapkan template HTML khusus untuk PDF
    template_path = 'pemeliharaan/struk_bensin_pdf.html'
    context = {'bensin': bensin}
    
    # Render HTML dengan data konteks
    template = get_template(template_path)
    html = template.render(context)
    
    # Buat buffer memory sementara
    buffer = BytesIO()
    
    # Konversi HTML ke PDF menggunakan xhtml2pdf
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    
    # Jika terjadi error saat konversi
    if pisa_status.err:
        return HttpResponse('Terjadi kesalahan saat membuat PDF <pre>' + html + '</pre>')
    
    # Ambil data PDF dari buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Buat response HTTP dengan file PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Struk_Bensin_{bensin.kendaraan.nomor_polisi}_{bensin.tanggal.strftime("%Y%m%d")}.pdf"'
    
    return response